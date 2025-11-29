"""
Excel Reader for Pic2Doc
Reads image filenames and captions from Excel files
"""

import openpyxl
from openpyxl.utils import column_index_from_string
from typing import List, Tuple
from pathlib import Path


class ExcelReader:
    """Reads image data from Excel files"""

    def __init__(self):
        """Initialize Excel reader"""
        pass

    def read_data(
        self,
        excel_path: str,
        filename_column: str = 'A',
        caption_columns: List[str] = None,
        caption_separator: str = ' - '
    ) -> List[Tuple[str, str]]:
        """
        Read image filenames and captions from Excel file

        Excel Format:
            - filename_column (default A): Filename without extension (e.g., "FIAS 21B 000001")
            - caption_columns (default ['I']): List of columns for multi-column captions

        Args:
            excel_path: Path to Excel file
            filename_column: Column letter for filenames (default 'A')
            caption_columns: List of column letters for captions (default ['I'])
            caption_separator: Separator for multi-column captions (default ' - ')

        Returns:
            List of tuples: (filename_without_ext, combined_caption)

        Raises:
            FileNotFoundError: If Excel file doesn't exist
            ValueError: If column structure is invalid
        """
        if caption_columns is None:
            caption_columns = ['I']

        excel_path = Path(excel_path)
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel-Datei nicht gefunden: {excel_path}")

        print(f"Lese Excel-Datei: {excel_path}")
        if len(caption_columns) > 1:
            print(f"  Bildunterschrift-Spalten: {', '.join(caption_columns)} (Trenner: '{caption_separator}')")

        # Load workbook
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # Convert column letters to indices
        filename_col_idx = column_index_from_string(filename_column)
        caption_col_indices = [column_index_from_string(col) for col in caption_columns]

        # Validate columns exist
        if filename_col_idx > ws.max_column:
            raise ValueError(f"Spalte {filename_column} nicht in Excel-Datei gefunden")
        for col, col_idx in zip(caption_columns, caption_col_indices):
            if col_idx > ws.max_column:
                raise ValueError(f"Spalte {col} nicht in Excel-Datei gefunden")

        data = []

        # Read from row 2 (skip header)
        start_row = 2
        max_col_idx = max([filename_col_idx] + caption_col_indices)

        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if len(row) < max_col_idx:
                continue

            # Get filename
            filename = row[filename_col_idx - 1]  # Convert to 0-based index

            # Get caption from multiple columns and combine
            caption_parts = []
            for col_idx in caption_col_indices:
                cell_value = row[col_idx - 1]  # Convert to 0-based index
                if cell_value:
                    caption_parts.append(str(cell_value).strip())

            caption = caption_separator.join(caption_parts) if caption_parts else ""

            if filename:
                filename_str = str(filename).strip()
                data.append((filename_str, caption))

        print(f"✓ {len(data)} Einträge gefunden")
        wb.close()

        return data

    def validate_structure(
        self,
        excel_path: str,
        filename_column: str = 'A',
        caption_column: str = 'I'
    ) -> bool:
        """
        Validate that Excel file has expected structure

        Args:
            excel_path: Path to Excel file
            filename_column: Expected filename column
            caption_column: Expected caption column

        Returns:
            True if structure is valid, False otherwise
        """
        try:
            excel_path = Path(excel_path)
            if not excel_path.exists():
                return False

            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            filename_col_idx = column_index_from_string(filename_column)
            caption_col_idx = column_index_from_string(caption_column)

            if filename_col_idx > ws.max_column or caption_col_idx > ws.max_column:
                wb.close()
                return False

            wb.close()
            return True
        except Exception:
            return False
